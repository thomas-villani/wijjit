"""Spreadsheet - edit a spreadsheet in the terminal, with a live chart.

A small "Excel in the console" app: load a workbook into an editable
``DataGrid``, tweak the numbers, save it back to disk, and watch a column chart
of per-product totals update alongside it.

Optional dependency
-------------------
If ``openpyxl`` is installed the app reads and writes a real ``budget.xlsx``.
If it is not, the app degrades gracefully to a stdlib-CSV ``budget.csv`` so the
demo still works end-to-end - only the on-disk format changes. Install the
spreadsheet extra with::

    pip install openpyxl

Features:
- Editable spreadsheet grid (arrow/Tab/Enter navigation, F2 or type to edit)
- Live column chart of each product's total across quarters
- Save to .xlsx (openpyxl) or .csv (fallback), and reload on next launch
- Add rows and reset to the sample data

Controls:
- Tab/Shift+Tab: Move focus between the grid and buttons
- In the grid: arrows/Tab/Enter to move, F2 or type to edit a cell
- Press Enter to commit a cell edit before clicking Save
- Ctrl+Q: Quit
"""

from __future__ import annotations

import csv
from pathlib import Path

from wijjit import Wijjit, render_template_string

try:
    import openpyxl

    HAVE_OPENPYXL = True
except ImportError:  # graceful fallback - see module docstring
    openpyxl = None  # type: ignore[assignment]
    HAVE_OPENPYXL = False

# Files live next to this script so the demo is self-contained regardless of the
# working directory it is launched from.
DATA_DIR = Path(__file__).resolve().parent
XLSX_PATH = DATA_DIR / "budget.xlsx"
CSV_PATH = DATA_DIR / "budget.csv"

# Column definitions shared by the grid, the chart, and the file I/O.
COLUMNS = [
    {"key": "product", "label": "Product", "width": 16},
    {"key": "q1", "label": "Q1", "width": 7},
    {"key": "q2", "label": "Q2", "width": 7},
    {"key": "q3", "label": "Q3", "width": 7},
    {"key": "q4", "label": "Q4", "width": 7},
]
HEADERS = [column["label"] for column in COLUMNS]

# Seed data used the first time the app runs (no file on disk yet).
SAMPLE_ROWS = [
    ["Widgets", "120", "135", "150", "160"],
    ["Gadgets", "90", "110", "95", "130"],
    ["Gizmos", "60", "75", "80", "70"],
    ["Doohickeys", "40", "55", "50", "65"],
]


def storage_path() -> Path:
    """Return the file the app reads and writes, based on available libraries."""
    return XLSX_PATH if HAVE_OPENPYXL else CSV_PATH


def storage_kind() -> str:
    """Return a short label for the active storage format ('xlsx' or 'csv')."""
    return "xlsx" if HAVE_OPENPYXL else "csv"


def load_rows() -> list[list[str]]:
    """Load rows from disk, falling back to the sample data.

    Returns
    -------
    list[list[str]]
        Data rows (header excluded), all cells as strings.
    """
    path = storage_path()
    if not path.exists():
        return [row[:] for row in SAMPLE_ROWS]

    if HAVE_OPENPYXL:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        rows = [
            ["" if cell is None else str(cell) for cell in record]
            for record in sheet.iter_rows(min_row=2, values_only=True)  # skip header
            if record is not None
        ]
    else:
        with path.open(newline="", encoding="utf-8") as handle:
            records = list(csv.reader(handle))
        rows = records[1:]  # skip header

    return rows or [row[:] for row in SAMPLE_ROWS]


def save_rows(rows: list[list[str]]) -> None:
    """Write rows (plus the header) to disk in the active format.

    Parameters
    ----------
    rows : list[list[str]]
        Data rows to persist.
    """
    path = storage_path()
    if HAVE_OPENPYXL:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Budget"
        sheet.append(HEADERS)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
    else:
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(HEADERS)
            writer.writerows(rows)


def compute_chart(rows: list[list[str]]) -> list[tuple[str, float]]:
    """Build (product, total) pairs summing the numeric quarter columns.

    Non-numeric cells are ignored so a half-typed edit never crashes the chart.

    Parameters
    ----------
    rows : list[list[str]]
        Current grid rows.

    Returns
    -------
    list[tuple[str, float]]
        One (label, value) pair per product row.
    """
    chart_data: list[tuple[str, float]] = []
    for row in rows:
        name = row[0] if row else ""
        total = 0.0
        for cell in row[1:]:
            try:
                total += float(cell)
            except (ValueError, TypeError):
                pass  # blank or non-numeric cell - skip it
        chart_data.append((name or "?", total))
    return chart_data


def current_rows() -> list[list[str]]:
    """Read the live grid contents (committed edits), falling back to state.

    DataGrid binding is one-directional (state -> grid at render), so edits live
    on the element until we read them back here.

    Returns
    -------
    list[list[str]]
        The grid's current data.
    """
    grid = app.get_element_by_id("sheet")
    if grid is not None:
        return grid.get_data()
    return app.state.get("sheet", [])


_initial_rows = load_rows()

app = Wijjit(
    initial_state={
        "sheet": _initial_rows,
        "chart_data": compute_chart(_initial_rows),
        "status": (
            f"Loaded {len(_initial_rows)} rows from {storage_path().name} "
            f"({storage_kind()} mode). Edit cells, then Save."
        ),
    }
)


@app.view("main", default=True)
def main_view():
    """Render the editable sheet next to a live totals chart."""
    return render_template_string(
        """
{% frame title="Terminal Spreadsheet" border="rounded" width="fill" height="fill" %}
  {% vstack spacing=1 padding=1 %}

    {% hstack spacing=2 %}

      {% frame title="Budget (editable)" border="single" width=54 height=16 %}
        {% datagrid id="sheet"
            data=sheet
            columns=columns
            width=50
            height=13
            editable=True
            border="none" %}
        {% enddatagrid %}
      {% endframe %}

      {% frame title="Totals by product" border="single" width=44 height=16 %}
        {% columnchart id="chart"
            data=chart_data
            width=40
            height=13
            column_width=3
            spacing=1
            show_axis=True
            border="none" %}
        {% endcolumnchart %}
      {% endframe %}

    {% endhstack %}

    {% hstack spacing=2 %}
      {% button action="save" %}Save{% endbutton %}
      {% button action="refresh_chart" %}Update Chart{% endbutton %}
      {% button action="add_row" %}Add Row{% endbutton %}
      {% button action="reset" %}Reset{% endbutton %}
    {% endhstack %}

    {% text %}{{ state.status }}{% endtext %}

  {% endvstack %}
{% endframe %}
        """,
        columns=COLUMNS,
        sheet=app.state["sheet"],
        chart_data=app.state["chart_data"],
    )


@app.on_action("refresh_chart")
def refresh_chart(event):
    """Recompute the chart from the current grid contents."""
    rows = current_rows()
    app.state["sheet"] = rows
    app.state["chart_data"] = compute_chart(rows)
    app.state["status"] = "Chart updated from the current grid."


@app.on_action("save")
def save(event):
    """Persist the grid to disk and refresh the chart."""
    rows = current_rows()
    app.state["sheet"] = rows
    app.state["chart_data"] = compute_chart(rows)
    try:
        save_rows(rows)
    except Exception as exc:  # keep the demo alive on a write error
        app.state["status"] = f"Save failed: {exc}"
        return
    app.state["status"] = f"Saved {len(rows)} rows to {storage_path().name}."


@app.on_action("add_row")
def add_row(event):
    """Append a blank product row to the sheet."""
    rows = current_rows() + [["New product", "0", "0", "0", "0"]]
    app.state["sheet"] = rows
    app.state["status"] = "Added a row. Edit it, then Save."


@app.on_action("reset")
def reset(event):
    """Restore the built-in sample data (does not write to disk)."""
    rows = [row[:] for row in SAMPLE_ROWS]
    app.state["sheet"] = rows
    app.state["chart_data"] = compute_chart(rows)
    app.state["status"] = "Reset to sample data (not yet saved)."


if __name__ == "__main__":
    app.run()
